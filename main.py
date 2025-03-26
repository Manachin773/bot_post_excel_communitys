from config import SessionLocal
from sqlalchemy import text
from GptAPi import GPT
from openpyxl import load_workbook
from datetime import datetime
import shutil
import random
import os

# === Configuración ===
plantilla_path = "Archivo de prueba post para Niyi.xlsx"
fecha_actual = datetime.today().strftime("%d/%m/%Y")
post_por_campaña = 5
carpeta_destino = "excel_campañas"

# Crear carpeta si no existe
if not os.path.exists(carpeta_destino):
    os.makedirs(carpeta_destino)

# === Conectar base de datos ===
with SessionLocal() as session:
    result = session.execute(text("SELECT campaign, commercial_services, residential_services, language FROM campaign"))
    rows = result.fetchall()

# === Generar Excel y Post por campaña ===
for row in rows:
    campaign_name, commercial_services, residential_services, lang = row
    campaign_key = campaign_name.lower().replace(" ", "_")
    print(f"\n📢 Procesando campaña: {campaign_name} [{campaign_key}]")

    # Crear copia nueva del Excel plantilla en la carpeta destino
    nuevo_excel = os.path.join(carpeta_destino, f"Post_{campaign_key}.xlsx")
    shutil.copy(plantilla_path, nuevo_excel)
    wb = load_workbook(nuevo_excel)
    ws = wb.active

    # Obtener cabeceras e índices de columnas
    headers = [cell.value for cell in ws[1]]
    idx_text = headers.index("Text") + 1
    idx_date = headers.index("Date") + 1
    idx_title = headers.index("Document title") + 1
    idx_comment = headers.index("First Comment Text") + 1

    # Detectar filas vacías
    filas_vacias = []
    for fila in range(2, ws.max_row + 1):
        if ws.cell(row=fila, column=idx_text).value in (None, ""):
            filas_vacias.append(fila)

    post_generados = 0

    for fila_objetivo in filas_vacias:
        if post_generados >= post_por_campaña:
            break

        opciones_servicios = []
        if commercial_services:
            opciones_servicios.append(commercial_services)
        if residential_services:
            opciones_servicios.append(residential_services)
        if not opciones_servicios:
            print(f"⚠️ No hay servicios en {campaign_name}")
            continue

        servicio_seleccionado = random.choice(random.choice(opciones_servicios).split(", "))
        print(f"📌 Servicio elegido: {servicio_seleccionado}")

        gpt = GPT({"service": servicio_seleccionado, "campaign": campaign_key, "lang": lang.lower()})

        # === Generar datos GPT según campaña ===
        if campaign_key == "osceola_fence_company":
            theme = gpt.theme_osceola()
            data = {
                "Text": gpt.copy_osceola(theme, 100),
                "Date": fecha_actual,
                "Document title": gpt.document_title_osceola(theme),
                "First Comment Text": gpt.firts_comment_osceola(theme, 50),
            }

        elif campaign_key == "quick_cleaning":
            theme = gpt.theme_quick_cleaning()
            data = {
                "Text": gpt.copy_quick_cleaning(theme, 100),
                "Date": fecha_actual,
                "Document title": gpt.document_title_quick_cleaning(theme),
                "First Comment Text": gpt.firts_comment_quick_cleaning(theme, 50),

            }

        elif campaign_key == "elite_chicago_spa":
            theme = gpt.theme_elite_spa()
            data = {
                "Text": gpt.copy_elite_spa(theme, 100),
                "Date": fecha_actual,
                "Document title": gpt.document_title_elite_spa(theme),
                "First Comment Text": gpt.firts_comment_elite_spa(theme, 50),
            }

        elif campaign_key == "lopez_&_lopez_abogados":
            theme = gpt.theme_lopez_abogados()
            data = {
                "Text": gpt.copy_lopez_abogados(theme, 100),
                "Date": fecha_actual,
                "Document title": gpt.document_title_lopez_abogados(theme),
                "First Comment Text": gpt.firts_comment_lopez_abogados(theme, 50),
            }

        elif campaign_key.startswith("botanica"):
            theme = gpt.theme_botanica()
            data = {
                "Text": gpt.copy_botanica(theme, 100),
                "Date": fecha_actual,
                "Document title": gpt.document_title_botanica(theme),
                "First Comment Text": gpt.firts_comment_botanica(theme, 50),    
            }

        else:
            print(f"⚠️ Campaña '{campaign_key}' no tiene métodos aún.")
            continue

        # === Llenar fila con datos ===
        ws.cell(row=fila_objetivo, column=idx_text).value = data["Text"]
        ws.cell(row=fila_objetivo, column=idx_date).value = data["Date"]
        ws.cell(row=fila_objetivo, column=idx_title).value = data["Document title"]
        ws.cell(row=fila_objetivo, column=idx_comment).value = data["First Comment Text"]
 

        post_generados += 1

    # === Limpiar celdas vacías extras (Metricool friendly) ===
    max_fila_valida = max(filas_vacias[:post_generados]) if post_generados else 1
    for fila in range(max_fila_valida + 1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=fila, column=col).value = None

    wb.save(nuevo_excel)
    print(f"✅ Excel guardado en carpeta '{carpeta_destino}': {nuevo_excel}")